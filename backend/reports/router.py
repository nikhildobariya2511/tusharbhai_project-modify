from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from models import Report, UploadedPDF
from schemas import ReportOut, ReportListItem,BatchDeleteRequest
from database import get_db
from utils import gen_report_no, compose_card_image
from auth.dependencies import get_current_user
import pandas as pd
import io
import os
import zipfile
from num2words import num2words
from openpyxl import load_workbook
from PIL import Image
from openpyxl_image_loader import SheetImageLoader
from typing import Optional, List, Dict
from datetime import datetime

router = APIRouter(
    prefix="/reports",
    tags=["Reports"],
    dependencies=[Depends(get_current_user)]
)

public_router = APIRouter(prefix="/public-report", tags=["Public Reports"])

UPLOAD_DIR = os.getenv("UPLOAD_DIR", "./uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ==========================================================
# Helper Function
# ==========================================================
def _reports_to_dataframe(reports: List[Report]) -> pd.DataFrame:
    rows = []
    for r in reports:
        rows.append({
            "report_no": r.report_no,
            "description": r.description,
            "shape_and_cut": r.shape_and_cut,
            "tot_est_weight": r.tot_est_weight,
            "color": r.color,
            "clarity": r.clarity,
            "style_number": r.style_number,
            "image_filename": r.image_filename,
            "comment": r.comment,
            "isecopy": bool(r.isecopy),
            "notice_image": bool(r.notice_image),
            "igi_logo": bool(getattr(r, "igi_logo", False)),
            "company_logo": r.company_logo,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return pd.DataFrame(rows)

# ==========================================================
# Export Backup
# ==========================================================
@router.post("/export-backup")
def export_backup(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    reports = db.query(Report).order_by(Report.created_at.asc()).all()
    df = _reports_to_dataframe(reports)

    excel_buf = io.BytesIO()
    with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="reports", index=False)
    excel_buf.seek(0)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("reports.xlsx", excel_buf.read())
        added = set()
        for r in reports:
            if r.image_filename:
                src = os.path.join(UPLOAD_DIR, r.image_filename)
                if os.path.exists(src) and r.image_filename not in added:
                    zf.write(src, arcname=f"images/{r.image_filename}")
                    added.add(r.image_filename)
            # include company logo files too (if present)
            if r.company_logo:
                logo_path = os.path.join(UPLOAD_DIR, "logo", r.company_logo)
                if os.path.exists(logo_path) and r.company_logo not in added:
                    zf.write(logo_path, arcname=f"logo/{r.company_logo}")
                    added.add(r.company_logo)
    zip_buf.seek(0)

    now_str = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"reports_backup_{now_str}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ==========================================================
# Import Backup
# ==========================================================
@router.post("/import-backup")
async def import_backup(
    file: UploadFile = File(...),
    overwrite: bool = Form(True),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Upload a .zip created by /reports/export-backup")
    blob = await file.read()
    buf = io.BytesIO(blob)

    try:
        with zipfile.ZipFile(buf, "r") as zf:
            if "reports.xlsx" not in zf.namelist():
                raise HTTPException(status_code=400, detail="reports.xlsx missing in zip")
            df = pd.read_excel(io.BytesIO(zf.read("reports.xlsx")), sheet_name="reports")
            df.columns = [c.strip() for c in df.columns]
            required = {
                "report_no","description","shape_and_cut","tot_est_weight",
                "color","clarity","style_number","image_filename",
                "comment","isecopy","created_at","notice_image","igi_logo"
            }
            missing = required - set(df.columns)
            if missing:
                raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(sorted(missing))}")

            os.makedirs(UPLOAD_DIR, exist_ok=True)
            image_members = {}
            logo_members = {}

            for name in zf.namelist():
                if name.startswith("images/") and len(name.split("/",1))==2:
                    image_members[name.split("/",1)[1]] = name
                if name.startswith("logo/") and len(name.split("/",1))==2:
                    logo_members[name.split("/",1)[1]] = name

            imported, updated, skipped = 0, 0, []

            def safe_str(val):
                return str(val).strip() if pd.notna(val) else None

            def parse_bool(val):
                if pd.isna(val):
                    return False
                if isinstance(val, bool):
                    return val
                return str(val).strip().lower() in {"true","1","yes","y","t"}

            for _, row in df.iterrows():
                report_no = safe_str(row["report_no"])
                if not report_no:
                    skipped.append("missing_report_no")
                    continue

                fields = {
                    "description": safe_str(row["description"]),
                    "shape_and_cut": safe_str(row["shape_and_cut"]),
                    "tot_est_weight": safe_str(row["tot_est_weight"]),
                    "color": safe_str(row["color"]),
                    "clarity": safe_str(row["clarity"]),
                    "style_number": safe_str(row["style_number"]),
                    "comment": safe_str(row["comment"]),
                }
                fields["isecopy"] = parse_bool(row["isecopy"])
                fields["notice_image"] = parse_bool(row["notice_image"])
                fields["igi_logo"] = parse_bool(row["igi_logo"])

                image_filename = safe_str(row["image_filename"])
                if image_filename and image_filename in image_members:
                    data = zf.read(image_members[image_filename])
                    with open(os.path.join(UPLOAD_DIR, image_filename), "wb") as f:
                        f.write(data)
                    fields["image_filename"] = image_filename
                else:
                    fields["image_filename"] = image_filename

                # company logo (if present in zip)
                company_logo_fn = safe_str(row.get("company_logo"))
                if company_logo_fn and company_logo_fn in logo_members:
                    logo_data = zf.read(logo_members[company_logo_fn])
                    logo_dir = os.path.join(UPLOAD_DIR, "logo")
                    os.makedirs(logo_dir, exist_ok=True)
                    with open(os.path.join(logo_dir, company_logo_fn), "wb") as f:
                        f.write(logo_data)
                    fields["company_logo"] = company_logo_fn
                else:
                    fields["company_logo"] = company_logo_fn

                existing = db.query(Report).filter(Report.report_no == report_no).first()
                if existing:
                    if overwrite:
                        for k, v in fields.items():
                            setattr(existing, k, v)
                        db.commit()
                        db.refresh(existing)
                        updated += 1
                    else:
                        skipped.append(report_no)
                else:
                    obj = Report(report_no=report_no, **fields)
                    db.add(obj)
                    db.commit()
                    db.refresh(obj)
                    imported += 1

        return {"imported": imported, "updated": updated, "skipped": skipped}
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Invalid zip file")

# ==========================================================
# New Route: Upload PDF ZIP (Seed PDF Data)
# ==========================================================
@router.post("/upload-pdf-zip")
async def upload_pdf_zip(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please upload a valid .zip file")

    pdf_dir = os.path.join(UPLOAD_DIR, "pdfs")
    os.makedirs(pdf_dir, exist_ok=True)

    blob = await file.read()
    buf = io.BytesIO(blob)

    try:
        with zipfile.ZipFile(buf, "r") as zf:
            pdf_files = [n for n in zf.namelist() if n.lower().endswith(".pdf")]
            if not pdf_files:
                raise HTTPException(status_code=400, detail="No PDF files found in zip")

            saved_files = []

            for name in pdf_files:
                base_name = os.path.basename(name)
                if not base_name:
                    continue

                report_no, _ = os.path.splitext(base_name)
                pdf_path = os.path.join(pdf_dir, f"{report_no}.pdf")

                # Save PDF
                with zf.open(name) as src, open(pdf_path, "wb") as dst:
                    shutil.copyfileobj(src, dst)

                # Insert DB row
                upload_log = UploadedPDF(
                    report_no=report_no,
                    filename=f"{report_no}.pdf",
                    uploaded_at=datetime.utcnow()
                )
                db.add(upload_log)
                saved_files.append(report_no)

            try:
                db.commit()
            except Exception as e:
                db.rollback()
                raise HTTPException(status_code=500, detail=f"DB Error: {str(e)}")

        return {"msg": f"{len(saved_files)} PDFs uploaded successfully", "reports": saved_files}

    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Invalid zip file")

    except Exception as e:
        print("\nZIP UPLOAD ERROR:", e)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@public_router.get("/{report_no}")
def get_public_report(report_no: str, db: Session = Depends(get_db)):
    """
    Return: uploads/pdfs/<report_no>.pdf (relative path)
    If PDF not found → return JSON report.
    """
    pdf_filename = f"{report_no}.pdf"
    pdf_path = os.path.join(UPLOAD_DIR, "pdfs", pdf_filename)

    # If the PDF exists → return relative path
    if os.path.exists(pdf_path):
        return {"pdf_path": f"uploads/pdfs/{pdf_filename}"}

    # fallback: return JSON from DB if PDF not found
    report = db.query(Report).filter(Report.report_no == report_no).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    return ReportOut.model_validate(report)


def extract_image_by_row(sheet, target_row):
    """
    Extract image anchored to the given Excel row.
    target_row = actual Excel row (1-based)
    """
    for img in getattr(sheet, "_images", []):
        try:
            anchor = img.anchor._from
            excel_row = anchor.row + 1  # convert 0-based to 1-based

            if excel_row == target_row:
                if hasattr(img, "_data"):
                    return Image.open(io.BytesIO(img._data()))
                if hasattr(img, "image"):
                    return img.image

        except Exception:
            continue

    return None


# -------------------------------------------------------------
#                MAIN UPLOAD ENDPOINT
# -------------------------------------------------------------
@router.post("/upload-xlsx")
def upload_xlsx(
    file: UploadFile = File(...),
    company_logo: UploadFile = File(None),
    diamond_type: Optional[str] = Form(None),
    comment: Optional[str] = Form(None),
    isecopy: bool = Form(False),
    notice_image: bool = Form(False),
    igi_logo: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):

    # ---------------------------------------------------------
    # 1) SAVE COMPANY LOGO (if uploaded)
    # ---------------------------------------------------------
    company_logo_filename = None

    if company_logo:
        ext = company_logo.filename.split(".")[-1].lower()
        if ext not in ["png", "jpg", "jpeg", "webp"]:
            raise HTTPException(400, "company_logo must be png/jpg/jpeg/webp")

        logo_dir = os.path.join(UPLOAD_DIR, "logo")
        os.makedirs(logo_dir, exist_ok=True)

        company_logo_filename = f"company_logo_{datetime.utcnow().timestamp()}.{ext}"
        logo_path = os.path.join(logo_dir, company_logo_filename)

        with open(logo_path, "wb") as f:
            f.write(company_logo.file.read())

    # ---------------------------------------------------------
    # 2) LOAD XLSX
    # ---------------------------------------------------------
    file_bytes = io.BytesIO(file.file.read())
    wb = load_workbook(file_bytes)
    sheet = wb["sheet1"] if "sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    df = pd.DataFrame(sheet.values)
    df.columns = [str(c).strip() for c in df.iloc[0]]
    df = df[1:]
    df = df.dropna(subset=["Jewelry Description", "Style Number"], how="all")
    df = df[df["Jewelry Description"].astype(str).str.strip() != ""]
    df = df.reset_index(drop=True)

    # remove rows containing (mandatory)
    df = df[~df.apply(lambda row: row.astype(str).str.contains(r"\(mandatory\)", case=False, na=False).any(), axis=1)]
    df = df.reset_index(drop=True)

    normalized_cols = {col.lower().strip(): col for col in df.columns}

    # ---------------------------------------------------------
    # 3) START PROCESSING ROWS
    # ---------------------------------------------------------
    reports = []
    skipped = []

    diamond_type_used = diamond_type.strip() if diamond_type else None

    for idx, row in df.iterrows():

        style = str(row.get("Style Number", "")).strip()
        if not style:
            continue

        # skip duplicates
        existing = db.query(Report).filter(Report.style_number == style).first()
        if existing:
            skipped.append(style)
            continue

        # comment field
        comment_val = comment.strip() if comment else None
        if not comment_val and "comment" in normalized_cols:
            col_name = normalized_cols["comment"]
            txt = str(row.get(col_name, "")).strip()
            if txt:
                comment_val = txt

        # diamond numbers
        try:
            num_diamonds_int = int(float(str(row.get("No Of Diamonds", 0)).strip()))
        except:
            num_diamonds_int = 0

        num_in_words = (
            num2words(num_diamonds_int, to="cardinal")
            .replace("-", " ")
            .capitalize()
        )

        diamonds_phrase = diamond_type_used if diamond_type_used else "Natural Diamonds"

        jewel_desc = str(row.get("Jewelry Description", "")).strip()

        # auto convert earrings
        if jewel_desc.lower() in ["earring", "ear rings", "ear-ring", "ear-rings"]:
            jewel_desc = "pair of earrings"

        desc = (
            f"One {row.get('Metal Color', '')} {jewel_desc}, "
            f"weighing in total {row.get('Gross Weight', '')}g, containing, "
            f"{num_in_words} ({num_diamonds_int}) {diamonds_phrase}"
        )

        shape = f"({num_diamonds_int}) {row.get('Shape', '')} Brilliant"
        tot = str(row.get("Diamond Weight", "")).strip()
        color = row.get("Color Criteria", "")
        clarity = row.get("Clarity Criteria", "")
        report_no = gen_report_no()

        # igi_logo (priority: form → column)
        igi_logo_val = bool(igi_logo)
        if not igi_logo_val:
            for key in ["igi_logo", "igi logo", "igi-logo"]:
                if key in normalized_cols:
                    val = row.get(normalized_cols[key])
                    if val:
                        igi_logo_val = str(val).strip().lower() in ["true", "1", "yes", "y"]
                    break

        # ---------------------------------------------------------
        # ⭐ 4) NEW FIX: GET IMAGE BY ANCHOR ROW
        # ---------------------------------------------------------
        excel_row = idx + 2  # Sheet row (header row is 1)
        img = extract_image_by_row(sheet, excel_row)

        image_filename = None
        if img:
            safe_style = style.replace(" ", "_")
            image_filename = f"{safe_style}.png"
            img_path = os.path.join(UPLOAD_DIR, image_filename)

            try:
                img.convert("RGBA").save(img_path, format="PNG", optimize=True)
            except:
                img.save(img_path, format="PNG")

        # ---------------------------------------------------------
        # 5) SAVE REPORT
        # ---------------------------------------------------------
        new_report = Report(
            report_no=report_no,
            description=desc,
            shape_and_cut=shape,
            tot_est_weight=tot,
            color=color,
            clarity=clarity,
            style_number=style,
            image_filename=image_filename,
            company_logo=company_logo_filename,
            comment=comment_val,
            notice_image=notice_image,
            isecopy=isecopy,
            igi_logo=igi_logo_val,
        )

        db.add(new_report)
        db.commit()
        db.refresh(new_report)

        reports.append(new_report)

    # ---------------------------------------------------------
    # 6) RETURN
    # ---------------------------------------------------------
    return {
        "uploaded": [ReportOut.model_validate(r) for r in reports],
        "skipped": skipped,
        "msg": f"{len(reports)} reports uploaded, {len(skipped)} skipped"
    }
# ==========================================================
# List Reports
# ==========================================================
@router.get("/", response_model=Dict[str, object])
def list_reports(
    q: Optional[str] = Query(None, description="Search report_no; partial match allowed"),
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
):
    query = db.query(Report)
    if q:
        query = query.filter(Report.report_no.ilike(f"%{q}%"))

    total = query.count()
    items = (
        query.order_by(Report.created_at.desc())
        .offset((page - 1) * size)
        .limit(size)
        .all()
    )

    result_items = [
        {"report_no": r.report_no, "style_number": r.style_number}
        for r in items
    ]

    return {
        "page": page,
        "size": size,
        "total": total,
        "items": result_items,
    }

# ==========================================================
# Update Report
# ==========================================================
@router.put("/{report_no}")
async def update_report(
    report_no: str,
    description: Optional[str] = Form(None),
    shape_and_cut: Optional[str] = Form(None),
    tot_est_weight: Optional[str] = Form(None),
    color: Optional[str] = Form(None),
    clarity: Optional[str] = Form(None),
    style_number: Optional[str] = Form(None),
    image_filename: Optional[str] = Form(None),
    notice_image: Optional[bool] = Form(None),
    comment: Optional[str] = Form(None),
    isecopy: Optional[bool] = Form(None),

    # ⬅ NEW: Company Logo Upload
    company_logo: UploadFile = File(None),

    # ⬅ NEW: replace/update the report image
    image: UploadFile = File(None),

    # ⬅ NEW: igi_logo boolean
    igi_logo: Optional[bool] = Form(None),

    db: Session = Depends(get_db),
):
    report = db.query(Report).filter(Report.report_no == report_no).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    # ------------------------------
    # UPDATE NORMAL REPORT IMAGE (replace)
    # ------------------------------
    if image:
        ext = image.filename.split(".")[-1].lower()
        if ext not in ["png", "jpg", "jpeg", "webp"]:
            raise HTTPException(400, "image must be PNG/JPG/JPEG/WEBP")

        # delete old image
        if report.image_filename:
            old = os.path.join(UPLOAD_DIR, report.image_filename)
            if os.path.exists(old):
                try:
                    os.remove(old)
                except:
                    pass

        new_filename = f"{report_no}_image.{ext}"
        new_path = os.path.join(UPLOAD_DIR, new_filename)

        content = await image.read()
        with open(new_path, "wb") as f:
            f.write(content)

        report.image_filename = new_filename

    # --------------------------
    # UPDATE COMPANY LOGO
    # --------------------------
    if company_logo:
        ext = company_logo.filename.split(".")[-1].lower()
        if ext not in ["png", "jpg", "jpeg", "webp"]:
            raise HTTPException(400, "company_logo must be PNG/JPG/JPEG/WEBP")

        logo_dir = os.path.join(UPLOAD_DIR, "logo")
        os.makedirs(logo_dir, exist_ok=True)

        # delete old company logo file (if exists)
        if report.company_logo:
            old_path = os.path.join(logo_dir, report.company_logo)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except:
                    pass

        new_logo_filename = f"company_logo_{report_no}_{int(datetime.utcnow().timestamp())}.{ext}"
        new_logo_path = os.path.join(logo_dir, new_logo_filename)

        content = await company_logo.read()
        with open(new_logo_path, "wb") as f:
            f.write(content)

        # update DB field
        report.company_logo = new_logo_filename

    # --------------------------
    # UPDATE OTHER FIELDS
    # --------------------------
    fields = {
        "description": description,
        "shape_and_cut": shape_and_cut,
        "tot_est_weight": tot_est_weight,
        "color": color,
        "clarity": clarity,
        "style_number": style_number,
        "image_filename": image_filename,
        "notice_image": notice_image,
        "comment": comment,
        "isecopy": isecopy,
        "igi_logo": igi_logo,
    }

    for key, value in fields.items():
        if value is not None:
            setattr(report, key, value)

    db.commit()
    db.refresh(report)

    return {
        "msg": "Report updated successfully",
        "report_no": report.report_no,
        "image_filename": report.image_filename,
        "company_logo": report.company_logo,
        "igi_logo": getattr(report, "igi_logo", False),
    }


# ==========================================================
# Delete Report
# ==========================================================
@router.delete("/{report_no}")
def delete_report(report_no: str, db: Session = Depends(get_db)):
    report = db.query(Report).filter(Report.report_no == report_no).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if report.image_filename:
        img_path = os.path.join(UPLOAD_DIR, report.image_filename)
        if os.path.exists(img_path):
            try:
                os.remove(img_path)
            except Exception:
                pass
    # delete company logo file (if exists)
    if report.company_logo:
        logo_path = os.path.join(UPLOAD_DIR, "logo", report.company_logo)
        if os.path.exists(logo_path):
            try:
                os.remove(logo_path)
            except Exception:
                pass

    db.delete(report)
    db.commit()
    return {"msg": f"Report {report_no} deleted"}


# ==========================================================
# Batch Delete Reports
# ==========================================================
@router.post("/batch-delete")
def batch_delete_reports(
    payload: BatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    report_nos = payload.report_no  # <-- singular key from schema

    deleted_count = 0
    failed = []

    for report_no in report_nos:
        try:
            report = (
                db.query(Report)
                .filter(Report.report_no == report_no)
                .first()
            )

            if not report:
                failed.append({"report_no": report_no, "error": "Not found"})
                continue

            # Delete image file
            if report.image_filename:
                img_path = os.path.join(UPLOAD_DIR, report.image_filename)
                if os.path.exists(img_path):
                    try:
                        os.remove(img_path)
                    except Exception:
                        pass

            # Delete company logo
            if report.company_logo:
                logo_path = os.path.join(UPLOAD_DIR, "logo", report.company_logo)
                if os.path.exists(logo_path):
                    try:
                        os.remove(logo_path)
                    except Exception:
                        pass

            db.delete(report)
            deleted_count += 1

        except Exception as e:
            failed.append({"report_no": report_no, "error": str(e)})

    db.commit()

    return {
        "msg": "Batch delete completed",
        "deleted": deleted_count,
        "failed": failed,
        "total": len(report_nos),
    }
